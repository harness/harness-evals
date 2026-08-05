#!/usr/bin/env python3
"""Build an Excel workbook from categorized conversation results + insights.

Reads one or more ``results.jsonl`` files and writes a multi-sheet workbook:

* ``All Results`` — every categorized row
* one sheet per dataset stem (e.g. ``module-coverage-200``, ``random-200``)
* ``Insights`` — category/module/env breakdowns, common prompt themes,
  highest-error patterns, and golden-readiness summary

Usage:
  python scripts/export_categorization_workbook.py \
      --results results/module-coverage-200/results.jsonl \
      --results results/random-200/results.jsonl \
      --output results/categorization-400.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULT_COLUMNS = [
    "dataset",
    "conversation_id",
    "original_prompt",
    "environment",
    "module",
    "usefulness",
    "quality",
    "golden_readiness",
    "final_category",
    "confidence",
    "goal_achievement",
    "resolution",
    "tool_use_quality",
    "score",
    "passed",
    "reasoning",
    "evidence",
    "canonical_file",
    "prompt_version",
]

PROMPT_THEMES: list[tuple[str, re.Pattern[str]]] = [
    ("pipeline_error_analysis", re.compile(r"analy[sz]e the error|pipeline execution|failed (pipeline|execution)|debug.*(pipeline|execution)", re.I)),
    ("create_pipeline", re.compile(r"\bcreate\b.*\bpipeline\b|\bnew pipeline\b", re.I)),
    ("update_pipeline", re.compile(r"\b(update|edit|modify|change)\b.*\bpipeline\b", re.I)),
    ("run_pipeline", re.compile(r"\b(run|execute|trigger)\b.*\bpipeline\b", re.I)),
    ("create_connector", re.compile(r"\bcreate\b.*\bconnector\b", re.I)),
    ("create_service", re.compile(r"\bcreate\b.*\bservice\b", re.I)),
    ("create_environment", re.compile(r"\bcreate\b.*\benvironment\b", re.I)),
    ("create_secret", re.compile(r"\bcreate\b.*\bsecret\b", re.I)),
    ("create_template", re.compile(r"\bcreate\b.*\btemplate\b|\btemplate\b.*\bplaceholder\b", re.I)),
    ("feature_flags", re.compile(r"feature flag|\bff\b|\btoggle\b", re.I)),
    ("delegates", re.compile(r"\bdelegate", re.I)),
    ("gitops", re.compile(r"\bgitops\b|\bargo\b|\bsync\b.*\bapp", re.I)),
    ("cost_ccm", re.compile(r"\bcost\b|\bccm\b|\bcloud spend\b|\banomal", re.I)),
    ("idp_catalog", re.compile(r"\bidp\b|\bscorecard\b|\bcatalog\b|\bsoftware template\b", re.I)),
    ("chaos", re.compile(r"\bchaos\b|\bfault inject", re.I)),
    ("security_sto", re.compile(r"\bvulnerabilit|\bscan\b|\bsto\b|\bsbom\b", re.I)),
    ("agents", re.compile(r"\bagent\b|\bmcp\b", re.I)),
    ("list_or_get", re.compile(r"\b(list|show|get|find|search)\b", re.I)),
    ("explain_how_to", re.compile(r"\bhow (do|can|to)\b|\bexplain\b|\bwhat is\b", re.I)),
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
SECTION_FONT = Font(bold=True, size=12)
WARN_FILL = PatternFill("solid", fgColor="F5B7B1")
GOOD_FILL = PatternFill("solid", fgColor="ABEBC6")
NI_FILL = PatternFill("solid", fgColor="F9E79F")


def _quality(row: dict[str, Any]) -> str:
    """Agent quality, with v2 ``needs_improvement`` mapped to ``good``."""
    quality = str(row.get("quality") or "").strip().lower()
    if quality:
        if quality == "needs_improvement":
            return "good"
        return quality
    final = str(row.get("final_category") or "unknown").strip().lower()
    if final == "needs_improvement":
        return "good"
    return final or "unknown"


def _readiness(row: dict[str, Any]) -> str:
    """Golden readiness: only ``ready`` or ``needs_rewrite``."""
    readiness = str(row.get("golden_readiness") or "").strip().lower()
    if readiness in {"ready", "needs_rewrite"}:
        return readiness
    final = str(row.get("final_category") or "").strip().lower()
    if final == "needs_improvement":
        return "needs_rewrite"
    if final == "good":
        return "ready"
    return "needs_rewrite"


def load_results(path: Path) -> list[dict[str, Any]]:
    dataset = path.parent.name
    rows: list[dict[str, Any]] = []
    with path.open() as handle:
        for line in handle:
            if not line.strip():
                continue
            row = json.loads(line)
            row["dataset"] = dataset
            evidence = row.get("evidence") or []
            if isinstance(evidence, list):
                row["evidence"] = " | ".join(str(item) for item in evidence)
            rows.append(row)
    return rows


def classify_prompt_theme(prompt: str) -> str:
    text = (prompt or "").strip()
    if not text:
        return "empty_prompt"
    for theme, pattern in PROMPT_THEMES:
        if pattern.search(text):
            return theme
    return "other"


def normalize_prompt_key(prompt: str) -> str:
    text = re.sub(r"\s+", " ", (prompt or "").strip().lower())
    text = re.sub(r"https?://\S+", "<url>", text)
    text = re.sub(r"\b[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\b", "<uuid>", text)
    text = re.sub(r"\b\d{10,}\b", "<id>", text)
    return text[:160]


def style_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx, name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:80]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def write_results_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    style_header(ws, RESULT_COLUMNS)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(RESULT_COLUMNS, start=1):
            value = row.get(key, "")
            if isinstance(value, float):
                value = round(value, 4)
            cell = ws.cell(row_idx, col_idx, value)
            if key == "quality":
                if value == "good":
                    cell.fill = GOOD_FILL
                elif value == "bad":
                    cell.fill = WARN_FILL
                elif value == "unclear":
                    cell.fill = NI_FILL
            if key == "golden_readiness":
                if value == "ready":
                    cell.fill = GOOD_FILL
                elif value == "needs_rewrite":
                    cell.fill = NI_FILL
            if key in {"original_prompt", "reasoning", "evidence"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)
    ws.row_dimensions[1].height = 22


def _write_section(ws, row: int, title: str, columns: list[str]) -> int:
    ws.cell(row, 1, title).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    for col_idx in range(2, len(columns) + 1):
        ws.cell(row, col_idx).fill = SECTION_FILL
    row += 1
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row, col_idx, name)
        cell.font = Font(bold=True)
    return row + 1


def write_insights_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Insights", 0)
    ws["A1"] = "Conversation Categorization Insights"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Total conversations: {len(rows)} | "
        f"Datasets: {', '.join(sorted({str(r.get('dataset')) for r in rows}))}"
    )

    for row in rows:
        row["_theme"] = classify_prompt_theme(str(row.get("original_prompt") or ""))
        row["_prompt_key"] = normalize_prompt_key(str(row.get("original_prompt") or ""))
        row["_quality"] = _quality(row)
        row["_readiness"] = _readiness(row)

    row_num = 4
    # Agent quality overview
    row_num = _write_section(ws, row_num, "1. Agent quality distribution", ["quality", "count", "pct"])
    cats = Counter(str(r["_quality"]) for r in rows)
    for category, count in cats.most_common():
        ws.cell(row_num, 1, category)
        ws.cell(row_num, 2, count)
        ws.cell(row_num, 3, round(100.0 * count / len(rows), 1) if rows else 0)
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "2. Golden readiness (independent of agent quality)",
        ["golden_readiness", "count", "pct", "guidance"],
    )
    readiness_counts = Counter(str(r["_readiness"]) for r in rows)
    readiness = [
        ("ready", readiness_counts.get("ready", 0), "Promote directly with org/project placeholders"),
        (
            "needs_rewrite",
            readiness_counts.get("needs_rewrite", 0),
            "Rewrite env-specific entity refs before promoting (can be good or bad)",
        ),
    ]
    for label, count, guidance in readiness:
        ws.cell(row_num, 1, label)
        ws.cell(row_num, 2, count)
        ws.cell(row_num, 3, round(100.0 * count / len(rows), 1) if rows else 0)
        ws.cell(row_num, 4, guidance)
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "3. Most common prompt themes",
        ["theme", "count", "pct", "good", "bad", "unclear", "needs_rewrite", "example_prompt"],
    )
    theme_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        theme_rows[str(row["_theme"])].append(row)
    for theme, theme_items in sorted(theme_rows.items(), key=lambda item: (-len(item[1]), item[0])):
        theme_cats = Counter(str(r["_quality"]) for r in theme_items)
        theme_ready = Counter(str(r["_readiness"]) for r in theme_items)
        example = next(
            (
                str(r.get("original_prompt") or "")[:180]
                for r in theme_items
                if r["_quality"] == "good"
            ),
            str(theme_items[0].get("original_prompt") or "")[:180],
        )
        ws.cell(row_num, 1, theme)
        ws.cell(row_num, 2, len(theme_items))
        ws.cell(row_num, 3, round(100.0 * len(theme_items) / len(rows), 1) if rows else 0)
        ws.cell(row_num, 4, theme_cats.get("good", 0))
        ws.cell(row_num, 5, theme_cats.get("bad", 0))
        ws.cell(row_num, 6, theme_cats.get("unclear", 0) + theme_cats.get("useless", 0))
        ws.cell(row_num, 7, theme_ready.get("needs_rewrite", 0))
        ws.cell(row_num, 8, example)
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "4. Exact/near-duplicate prompts (normalized)",
        ["normalized_prompt", "count", "dominant_quality", "example_conversation_id"],
    )
    prompt_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = str(row["_prompt_key"])
        if key:
            prompt_groups[key].append(row)
    for key, group in sorted(prompt_groups.items(), key=lambda item: (-len(item[1]), item[0]))[:25]:
        if len(group) < 2:
            continue
        dominant = Counter(str(r["_quality"]) for r in group).most_common(1)[0][0]
        ws.cell(row_num, 1, key)
        ws.cell(row_num, 2, len(group))
        ws.cell(row_num, 3, dominant)
        ws.cell(row_num, 4, group[0].get("conversation_id"))
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "5. Most errored themes (bad + low goal achievement)",
        ["theme", "bad_count", "avg_goal_achievement", "avg_score", "example_prompt", "example_reasoning"],
    )
    errored = [r for r in rows if r["_quality"] == "bad" or float(r.get("goal_achievement") or 1) < 0.5]
    errored_by_theme: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in errored:
        errored_by_theme[str(row["_theme"])].append(row)
    if not errored_by_theme:
        ws.cell(row_num, 1, "(no bad / low-goal conversations in this batch)")
        row_num += 1
    else:
        for theme, items in sorted(errored_by_theme.items(), key=lambda item: (-len(item[1]), item[0])):
            goals = [float(r["goal_achievement"]) for r in items if r.get("goal_achievement") is not None]
            scores = [float(r["score"]) for r in items if r.get("score") is not None]
            sample = items[0]
            ws.cell(row_num, 1, theme)
            ws.cell(row_num, 2, sum(1 for r in items if r["_quality"] == "bad"))
            ws.cell(row_num, 3, round(sum(goals) / len(goals), 3) if goals else "")
            ws.cell(row_num, 4, round(sum(scores) / len(scores), 3) if scores else "")
            ws.cell(row_num, 5, str(sample.get("original_prompt") or "")[:180])
            ws.cell(row_num, 6, str(sample.get("reasoning") or "")[:220])
            row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "6. Agent quality by module",
        ["module", "total", "good", "bad", "unclear", "useless", "needs_rewrite"],
    )
    by_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_module[str(row.get("module") or "unknown")].append(row)
    for module, items in sorted(by_module.items(), key=lambda item: (-len(item[1]), item[0])):
        module_cats = Counter(str(r["_quality"]) for r in items)
        module_ready = Counter(str(r["_readiness"]) for r in items)
        ws.cell(row_num, 1, module)
        ws.cell(row_num, 2, len(items))
        for col_idx, cat in enumerate(["good", "bad", "unclear", "useless"], start=3):
            ws.cell(row_num, col_idx, module_cats.get(cat, 0))
        ws.cell(row_num, 7, module_ready.get("needs_rewrite", 0))
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "7. Agent quality by environment",
        ["environment", "total", "good", "bad", "unclear", "useless", "needs_rewrite"],
    )
    by_env: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_env[str(row.get("environment") or "unknown")].append(row)
    for env, items in sorted(by_env.items(), key=lambda item: (-len(item[1]), item[0])):
        env_cats = Counter(str(r["_quality"]) for r in items)
        env_ready = Counter(str(r["_readiness"]) for r in items)
        ws.cell(row_num, 1, env)
        ws.cell(row_num, 2, len(items))
        for col_idx, cat in enumerate(["good", "bad", "unclear", "useless"], start=3):
            ws.cell(row_num, col_idx, env_cats.get(cat, 0))
        ws.cell(row_num, 7, env_ready.get("needs_rewrite", 0))
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "8. Top ready + good candidates (direct golden promotion)",
        ["dataset", "module", "confidence", "prompt", "conversation_id"],
    )
    goods = sorted(
        (r for r in rows if r["_quality"] == "good" and r["_readiness"] == "ready"),
        key=lambda r: (-float(r.get("confidence") or 0), str(r.get("module") or "")),
    )[:30]
    if not goods:
        ws.cell(row_num, 1, "(none categorized as good + ready)")
        row_num += 1
    else:
        for item in goods:
            ws.cell(row_num, 1, item.get("dataset"))
            ws.cell(row_num, 2, item.get("module"))
            ws.cell(row_num, 3, item.get("confidence"))
            ws.cell(row_num, 4, str(item.get("original_prompt") or "")[:200])
            ws.cell(row_num, 5, item.get("conversation_id"))
            row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "9. Top needs_rewrite candidates (any agent quality)",
        ["dataset", "module", "quality", "theme", "prompt", "why_not_portable"],
    )
    needs = sorted(
        (r for r in rows if r["_readiness"] == "needs_rewrite"),
        key=lambda r: (-float(r.get("score") or 0), -float(r.get("confidence") or 0)),
    )[:30]
    for item in needs:
        ws.cell(row_num, 1, item.get("dataset"))
        ws.cell(row_num, 2, item.get("module"))
        ws.cell(row_num, 3, item.get("_quality"))
        ws.cell(row_num, 4, item.get("_theme"))
        ws.cell(row_num, 5, str(item.get("original_prompt") or "")[:200])
        ws.cell(row_num, 6, str(item.get("reasoning") or "")[:220])
        row_num += 1

    row_num += 1
    row_num = _write_section(
        ws,
        row_num,
        "10. Dataset comparison",
        ["dataset", "total", "good", "bad", "unclear", "useless", "needs_rewrite", "avg_confidence"],
    )
    by_dataset: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_dataset[str(row.get("dataset") or "unknown")].append(row)
    for dataset, items in sorted(by_dataset.items()):
        dataset_cats = Counter(str(r["_quality"]) for r in items)
        dataset_ready = Counter(str(r["_readiness"]) for r in items)
        confs = [float(r["confidence"]) for r in items if r.get("confidence") is not None]
        ws.cell(row_num, 1, dataset)
        ws.cell(row_num, 2, len(items))
        for col_idx, cat in enumerate(["good", "bad", "unclear", "useless"], start=3):
            ws.cell(row_num, col_idx, dataset_cats.get(cat, 0))
        ws.cell(row_num, 7, dataset_ready.get("needs_rewrite", 0))
        ws.cell(row_num, 8, round(sum(confs) / len(confs), 3) if confs else "")
        row_num += 1

    autosize(ws, max_width=80)


def build_workbook(result_paths: list[Path], output: Path) -> dict[str, Any]:
    all_rows: list[dict[str, Any]] = []
    by_dataset: dict[str, list[dict[str, Any]]] = {}
    for path in result_paths:
        rows = load_results(path)
        by_dataset[path.parent.name] = rows
        all_rows.extend(rows)

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    write_insights_sheet(wb, all_rows)
    write_results_sheet(wb, "All Results", all_rows)
    for dataset, rows in sorted(by_dataset.items()):
        write_results_sheet(wb, dataset, rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    cats = Counter(_quality(r) for r in all_rows)
    readiness = Counter(_readiness(r) for r in all_rows)
    themes = Counter(classify_prompt_theme(str(r.get("original_prompt") or "")) for r in all_rows)
    return {
        "total": len(all_rows),
        "qualities": dict(cats),
        "golden_readiness": dict(readiness),
        "categories": dict(cats),
        "top_themes": themes.most_common(8),
        "output": str(output),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--results", type=Path, action="append", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    for path in args.results:
        if not path.is_file():
            raise SystemExit(f"Missing results file: {path}")
    summary = build_workbook(args.results, args.output)
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
